import os
import chromadb
from chromadb.utils import embedding_functions
import pypdf
import openpyxl
from docx import Document
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize ChromaDB
# We use a persistent client so data is saved to disk
DB_DIR = os.path.join(os.path.dirname(__file__), "chroma_db")
client = chromadb.PersistentClient(path=DB_DIR)

# Use Google Generative AI embeddings (since we use Gemini)
# Note: This requires GOOGLE_API_KEY environment variable to be set
# For simplicity in this PoC, we can also use a default sentence-transformer if preferred,
# but using Gemini embeddings often aligns better.
# Let's try to use the default all-MiniLM-L6-v2 for now as it doesn't require an extra API call per chunk,
# making it faster and free for local testing.
embedding_func = embedding_functions.SentenceTransformerEmbeddingFunction(model_name="all-MiniLM-L6-v2")

collection = client.get_or_create_collection(
    name="adda_knowledge_base",
    embedding_function=embedding_func
)

def extract_text_from_pdf(file_path):
    text = ""
    try:
        reader = pypdf.PdfReader(file_path)
        for page in reader.pages:
            text += page.extract_text() + "\n"
    except Exception as e:
        logger.error(f"Error reading PDF {file_path}: {e}")
    return text

def extract_text_from_excel(file_path):
    text = ""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet in workbook.sheetnames:
            text += f"Sheet: {sheet}\n"
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) for cell in row if cell is not None])
                text += row_text + "\n"
    except Exception as e:
        logger.error(f"Error reading Excel {file_path}: {e}")
    return text

def extract_text_from_docx(file_path):
    text = ""
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        logger.error(f"Error reading Docx {file_path}: {e}")
    return text

def chunk_text(text, chunk_size=1000, overlap=100):
    chunks = []
    start = 0
    text_len = len(text)
    
    while start < text_len:
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap
        
    return chunks

def add_document(file_path, original_filename, category=None):
    """
    Reads a file, chunks the text, and adds it to the vector database.
    
    Args:
        file_path: Path to the file to process
        original_filename: Original name of the file
        category: Optional category tag (e.g., "roles", "levels", "rules")
                  Determined by folder structure if not provided
    """
    ext = os.path.splitext(original_filename)[1].lower()
    text = ""
    
    # Auto-detect category from file path if not provided
    if category is None:
        if "/roles/" in file_path or "\\roles\\" in file_path:
            category = "roles"
        elif "/levels/" in file_path or "\\levels\\" in file_path:
            category = "levels"
        elif "/rules/" in file_path or "\\rules\\" in file_path:
            category = "rules"
        else:
            category = "general"
    
    if ext == ".pdf":
        text = extract_text_from_pdf(file_path)
    elif ext in [".xlsx", ".xls"]:
        text = extract_text_from_excel(file_path)
    elif ext in [".docx", ".doc"]:
        text = extract_text_from_docx(file_path)
    else:
        # Try text fallback
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
        except:
            logger.error(f"Unsupported file format: {ext}")
            return False

    if not text:
        logger.warning(f"No text extracted from {original_filename}")
        return False

    chunks = chunk_text(text)
    
    ids = [f"{original_filename}_{i}" for i in range(len(chunks))]
    metadatas = [{"source": original_filename, "chunk_id": i, "category": category} for i in range(len(chunks))]
    
    try:
        collection.add(
            documents=chunks,
            ids=ids,
            metadatas=metadatas
        )
        logger.info(f"Successfully added {len(chunks)} chunks from {original_filename} (category: {category})")
        return True
    except Exception as e:
        logger.error(f"Error adding to ChromaDB: {e}")
        return False

def query_knowledge_base(query, n_results=3, category=None):
    """
    Queries the vector database for relevant chunks.
    
    Args:
        query: The search query
        n_results: Number of results to return
        category: Optional category filter (e.g., "roles", "levels", "rules")
    
    Returns:
        Dict with documents and metadatas
    """
    try:
        # First try with category filter if provided
        if category:
            try:
                results = collection.query(
                    query_texts=[query],
                    n_results=n_results,
                    where={"category": category}
                )
                documents = results['documents'][0] if results['documents'] else []
                if documents:
                    metadatas = results['metadatas'][0] if results['metadatas'] else []
                    return {
                        "documents": documents,
                        "metadatas": metadatas
                    }
            except Exception as e:
                logger.warning(f"Category search failed, falling back to general search: {e}")
        
        # Fallback: search without category filter
        results = collection.query(
            query_texts=[query],
            n_results=n_results
        )
        
        # Flatten results (Chroma returns list of lists)
        documents = results['documents'][0] if results['documents'] else []
        metadatas = results['metadatas'][0] if results['metadatas'] else []
        
        return {
            "documents": documents,
            "metadatas": metadatas
        }
    except Exception as e:
        logger.error(f"Error querying ChromaDB: {e}")
        return {"documents": [], "metadatas": []}

def query_by_category(query, category, n_results=5):
    """
    Convenience function to query a specific category.
    Used by agent system to get context from specific document types.
    
    Args:
        query: The search query
        category: Category to search in ("roles", "levels", "rules", "general")
        n_results: Number of results to return
    
    Returns:
        List of document strings
    """
    result = query_knowledge_base(query, n_results=n_results, category=category)
    return result['documents']

def get_full_text_from_document(filename):
    """
    Retrieves all chunks from a specific document and reconstructs the full text.
    Useful for document analysis.
    """
    try:
        # Get all chunks from this source
        results = collection.get(
            where={"source": filename}
        )
        
        if not results or not results['documents']:
            return ""
        
        # Sort by chunk_id to preserve order
        chunks_with_meta = list(zip(results['documents'], results['metadatas']))
        chunks_with_meta.sort(key=lambda x: x[1].get('chunk_id', 0))
        
        # Reconstruct text
        full_text = " ".join([chunk for chunk, _ in chunks_with_meta])
        return full_text
        
    except Exception as e:
        logger.error(f"Error retrieving document {filename}: {e}")
        return ""


# =============================================================================
# MANIFEST-BASED QUERIES (Agent Scoping)
# =============================================================================

def query_with_manifest(query, manifest, n_results=5):
    """
    Query the knowledge base with manifest-based filtering.
    
    This is the main function for agent-scoped RAG queries.
    It ensures that each step only "sees" the documents it's allowed to access.
    
    Args:
        query: The search query
        manifest: Dict with 'allowed_categories' and 'specific_files' lists
        n_results: Number of results to return
    
    Returns:
        Dict with 'documents' and 'metadatas' lists
    """
    allowed_categories = manifest.get("allowed_categories", [])
    specific_files = manifest.get("specific_files", [])
    
    logger.info(f"Query with manifest - categories: {allowed_categories}, files: {specific_files}")
    
    all_documents = []
    all_metadatas = []
    
    # Query by allowed categories
    if allowed_categories:
        for category in allowed_categories:
            try:
                results = collection.query(
                    query_texts=[query],
                    n_results=n_results,
                    where={"category": category}
                )
                
                if results['documents'] and results['documents'][0]:
                    all_documents.extend(results['documents'][0])
                    if results['metadatas'] and results['metadatas'][0]:
                        all_metadatas.extend(results['metadatas'][0])
                        
            except Exception as e:
                logger.warning(f"Error querying category '{category}': {e}")
    
    # Query by specific files
    if specific_files:
        for filename in specific_files:
            try:
                results = collection.query(
                    query_texts=[query],
                    n_results=n_results,
                    where={"source": filename}
                )
                
                if results['documents'] and results['documents'][0]:
                    # Avoid duplicates
                    for doc, meta in zip(results['documents'][0], results['metadatas'][0] or [{}] * len(results['documents'][0])):
                        if doc not in all_documents:
                            all_documents.append(doc)
                            all_metadatas.append(meta)
                            
            except Exception as e:
                logger.warning(f"Error querying file '{filename}': {e}")
    
    # If no manifest filters, fall back to general query
    if not allowed_categories and not specific_files:
        logger.info("No manifest filters, performing general query")
        return query_knowledge_base(query, n_results=n_results)
    
    # Limit results
    if len(all_documents) > n_results:
        all_documents = all_documents[:n_results]
        all_metadatas = all_metadatas[:n_results]
    
    logger.info(f"Manifest query returned {len(all_documents)} documents")
    
    return {
        "documents": all_documents,
        "metadatas": all_metadatas
    }


def list_all_documents():
    """
    List all indexed documents with their metadata.
    
    Returns:
        List of dicts with 'filename', 'category', 'chunk_count'
    """
    try:
        # Get all documents
        results = collection.get()
        
        if not results or not results['metadatas']:
            return []
        
        # Group by source file
        files = {}
        for meta in results['metadatas']:
            source = meta.get('source', 'unknown')
            category = meta.get('category', 'general')
            
            if source not in files:
                files[source] = {
                    'filename': source,
                    'category': category,
                    'chunk_count': 0
                }
            files[source]['chunk_count'] += 1
        
        return list(files.values())
        
    except Exception as e:
        logger.error(f"Error listing documents: {e}")
        return []


def remove_document(filename):
    """
    Remove all chunks associated with a specific filename.
    
    Args:
        filename: The source filename to remove
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Get IDs of all chunks from this source
        results = collection.get(
            where={"source": filename}
        )
        
        if not results or not results['ids']:
            logger.warning(f"No documents found with filename: {filename}")
            return False
        
        # Delete by IDs
        collection.delete(ids=results['ids'])
        logger.info(f"Removed {len(results['ids'])} chunks from {filename}")
        return True
        
    except Exception as e:
        logger.error(f"Error removing document {filename}: {e}")
        return False


def get_collection_stats():
    """
    Get statistics about the knowledge base.
    
    Returns:
        Dict with collection statistics
    """
    try:
        count = collection.count()
        docs = list_all_documents()
        
        categories = {}
        for doc in docs:
            cat = doc['category']
            if cat not in categories:
                categories[cat] = 0
            categories[cat] += doc['chunk_count']
        
        return {
            "total_chunks": count,
            "total_documents": len(docs),
            "categories": categories,
            "documents": docs
        }
        
    except Exception as e:
        logger.error(f"Error getting stats: {e}")
        return {
            "total_chunks": 0,
            "total_documents": 0,
            "categories": {},
            "documents": []
        }
